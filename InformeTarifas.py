import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
import numpy as np
import zipfile
import io

st.title("Generación de informes (tarifas e informe Dane)")

# Subir archivos (un solo botón para todos)
uploaded_files = st.file_uploader(
    "Subir archivos (TC1.csv, TC2.xlsx, AP.xlsx, Divipola.xlsx, Bitacora.xlsx, Informe_Tarifas_def.xlsx)",
    type=["csv", "xlsx"],
    accept_multiple_files=True
)

# Diccionario para almacenar los archivos subidos
file_dict = {
    "TC1": None,
    "TC2": None,
    "AP": None,
    "DIVIPOLA": None,
    "BITACORA": None,
    "INFORME_TARIFAS_DEF": None
}

# Asociar cada archivo subido a su clave correspondiente
if uploaded_files:
    for file in uploaded_files:
        name_upper = file.name.upper()
        if "TC1" in name_upper:
            file_dict["TC1"] = file
        elif "TC2" in name_upper:
            file_dict["TC2"] = file
        elif "AP" in name_upper:
            file_dict["AP"] = file
        elif "DIVIPOLA" in name_upper:
            file_dict["DIVIPOLA"] = file
        elif "BITACORA" in name_upper:
            file_dict["BITACORA"] = file
        elif "INFORME_TARIFAS_DEF" in name_upper:
            file_dict["INFORME_TARIFAS_DEF"] = file

# Verificar que se hayan cargado todos los archivos
if all(file_dict.values()):
    try:
        # Leer archivos
        tc1 = pd.read_csv(file_dict["TC1"], low_memory=False)
        tc2 = pd.read_excel(file_dict["TC2"])
        ap = pd.read_excel(file_dict["AP"], sheet_name="TABLA TARIFAS", header=3)
        # Excluir filas que contienen "Total general"
        ap = ap[~ap.iloc[:, 0].astype(str).str.contains("Total general", na=False)]
        divipola = pd.read_excel(file_dict["DIVIPOLA"])
        bitacora = pd.read_excel(file_dict["BITACORA"])

        # Aplicar filtro en TC1
        id_comercializador_col = 'ID COMERCIALIZADOR'
        niu_col = 'NIU'
        if id_comercializador_col in tc1.columns and niu_col in tc1.columns:
            tc1_filtrado = tc1[tc1[id_comercializador_col] == 23442]
            count_nius_tc1 = tc1_filtrado[niu_col].nunique()
            st.write(f"Número de NIUs en TC1 después de filtrar: {count_nius_tc1}")
        else:
            st.error("Las columnas esperadas no están en TC1.")

        # Validación de TC2 (NIUs y Tarifas)
        if niu_col in tc2.columns:
            tc2_sin_duplicados = tc2.drop_duplicates(subset=niu_col)
            count_nius_tc2 = tc2_sin_duplicados[niu_col].nunique()
            st.write(f"Número de NIUs en TC2 después de eliminar duplicados: {count_nius_tc2}")

            if count_nius_tc1 == count_nius_tc2 - 1:
                st.success("✅ El número de NIUs en TC2 coincide con el valor esperado.")
            else:
                st.error("❌ El número de NIUs en TC2 no coincide con el valor esperado. Verifica los archivos.")

            if 'Tipo de Tarifa' in tc2.columns:
                duplicated_nius = tc2[tc2.duplicated(subset='NIU', keep=False)]
                different_tarifas = duplicated_nius.groupby('NIU')['Tipo de Tarifa'].nunique()
                nius_with_different_tarifas = different_tarifas[different_tarifas > 1]
                if not nius_with_different_tarifas.empty:
                    st.error("❌ Hay NIUs con diferentes tipos de tarifa. Revisa los datos.")
                    niu_different_tarifa_df = duplicated_nius[
                        duplicated_nius['NIU'].isin(nius_with_different_tarifas.index)
                    ]
                    st.write("### NIUs con tipo de tarifa diferente:")
                    st.dataframe(niu_different_tarifa_df[['NIU', 'Tipo de Tarifa']])
                else:
                    st.success("✅ Todos los NIUs tienen el mismo tipo de tarifa.")
        else:
            st.error("❌ Las columnas esperadas no están en TC2.")

    except Exception as e:
        st.error(f"Ocurrió un error al procesar los archivos: {e}")

    # Generación de Tabla de Tarifas
    required_columns = [
        'NIU', 'CODIGO DANE (NIU)', 'ESTRATO', 'UBICACION',
        'NIVEL DE TENSION', 'PORCENTAJE PROPIEDAD DEL ACTIVO', 'CODIGO AREA ESPECIAL'
    ]
    if all(col in tc1_filtrado.columns for col in required_columns):
        Tarifas = tc1_filtrado[required_columns].copy()
        Tarifas.columns = [
            'NIU', 'DIVIPOLA', 'ESTRATO', 'UBICACION', 
            'NIVEL DE TENSION', 'CARGA DE INVERSION', 'ZE'
        ]
        Tarifas['ESTRATO'] = Tarifas['ESTRATO'].replace({7: 'I', 8: 'C', 9: 'O', 11: 'AP'})
        Tarifas['UBICACION'] = Tarifas['UBICACION'].replace({1: 'R', 2: 'U'})
        Tarifas['CARGA DE INVERSION'] = Tarifas['CARGA DE INVERSION'].replace({101: 0})

        # Combinar con divipola para obtener el nombre del municipio
        divipola.columns = divipola.columns.str.strip()
        Tarifas = Tarifas.merge(
            divipola[['Código DIVIPOLA', 'Nombre Municipio']],
            left_on='DIVIPOLA', right_on='Código DIVIPOLA', how='left'
        )
        Tarifas = Tarifas.rename(columns={'Nombre Municipio': 'MUNICIPIO'})
        Tarifas = Tarifas.drop(columns=['Código DIVIPOLA'])

        # Crear tabla dinámica a partir de TC2
        pivot_table = pd.pivot_table(
            tc2, index='NIU',
            values=['Consumo Usuario (kWh)', 'Valor Facturación por Consumo Usuario'],
            aggfunc='sum'
        )
        pivot_table.reset_index(inplace=True)
        tblDinamicaTc2 = pivot_table[['NIU', 'Consumo Usuario (kWh)', 'Valor Facturación por Consumo Usuario']]

        # Asegurarse de que los NIU sean cadenas sin espacios en blanco
        Tarifas['NIU'] = Tarifas['NIU'].astype(str).str.strip()
        tblDinamicaTc2['NIU'] = tblDinamicaTc2['NIU'].astype(str).str.strip()
        tc2_sin_duplicados['NIU'] = tc2_sin_duplicados['NIU'].astype(str).str.strip()

        tblDinamicaTc2 = tblDinamicaTc2.merge(
            tc2_sin_duplicados[['NIU', 'Tipo de Tarifa']],
            on='NIU', how='left'
        )
        Tarifas = Tarifas.merge(tblDinamicaTc2, on='NIU', how='left')
        Tarifas['Tipo de Tarifa'] = Tarifas['Tipo de Tarifa'].replace({1: 'R', 2: 'NR'})

        Tarifas = Tarifas[[
            'NIU', 'ESTRATO', 'Tipo de Tarifa', 'Consumo Usuario (kWh)',
            'Valor Facturación por Consumo Usuario', 'UBICACION',
            'DIVIPOLA', 'MUNICIPIO', 'NIVEL DE TENSION',
            'CARGA DE INVERSION', 'ZE'
        ]]
        Tarifas = Tarifas.rename(columns={
            'Tipo de Tarifa': 'TIPO TARIFA',
            'Consumo Usuario (kWh)': 'CONSUMO',
            'Valor Facturación por Consumo Usuario': 'FACTURACION CONSUMO'
        })

        # Añadir cliente de otro mercado si existe
        niu_filtrado = tc2[(tc2['NIU'] == 898352932) | (tc2['NIU'] == 18124198)]
        if not niu_filtrado.empty:
            consumo_usuario = niu_filtrado['Consumo Usuario (kWh)'].values[0]
            valor_facturacion = niu_filtrado['Valor Facturación por Consumo Usuario'].values[0]
            nueva_fila = pd.DataFrame({
                'NIU': [898352932],
                'ESTRATO': ['I'],
                'TIPO TARIFA': ['NR'],
                'CONSUMO': [consumo_usuario],
                'FACTURACION CONSUMO': [valor_facturacion],
                'UBICACION': ['U'],
                'DIVIPOLA': [13001000],
                'MUNICIPIO': ['CARTAGENA'],
                'NIVEL DE TENSION': [2],
                'CARGA DE INVERSION': [0],
                'ZE': [0]
            })
            Tarifas = pd.concat([Tarifas, nueva_fila], ignore_index=True)

        # Eliminar NIUs que contienen 'CAL'
        Tarifas['NIU'] = Tarifas['NIU'].astype(str).fillna('')
        Tarifas = Tarifas[~Tarifas['NIU'].str.contains('CAL')]

        # Procesar el archivo AP
        ap['producto'] = ap['producto'].astype(str).str.strip()
        if ap['producto'].eq('').any():
            st.error("❌ El archivo AP contiene productos vacíos. Corrige los datos.")
            st.stop()
        else:
            st.success("✅ Validación exitosa: No hay productos vacíos en el archivo AP.")

        ap['tipo de tarifa'] = ap['tipo de tarifa'].replace({1: 'R', 2: 'NR'})
        ap['estrato'] = ap['estrato'].replace({11: 'AP'})
        ap = ap[ap['estrato'] == 'AP']

        tarifas_val = Tarifas[Tarifas['ESTRATO'] == 'AP']
        nius_archivo_ap = set(ap['producto'].astype(str).str.strip())
        nius_tarifas_ap = set(tarifas_val['NIU'].astype(str).str.strip())
        niu_faltantes_en_ap = nius_tarifas_ap - nius_archivo_ap
        niu_faltantes_en_tarifas = nius_archivo_ap - nius_tarifas_ap

        if niu_faltantes_en_ap:
            st.error(f"❌ NIU en Tarifas (AP) que no están en archivo AP: {niu_faltantes_en_ap}")
            st.stop()
        if niu_faltantes_en_tarifas:
            st.error(f"❌ NIU en archivo AP (AP) que no están en Tarifas: {niu_faltantes_en_tarifas}")
            st.stop()
        else:
            st.success("✅ Validación exitosa: se puede hacer cruce de AP con tarifas.")

        Tarifas = Tarifas.merge(
            ap[['producto', 'Suma de consumo', 'Suma de facturacion consumo', 'tipo de tarifa']],
            left_on='NIU', right_on='producto', how='left'
        )
        Tarifas.loc[Tarifas['Suma de consumo'] > 0, 'CONSUMO'] = Tarifas['Suma de consumo']
        Tarifas.loc[
            (Tarifas['Suma de facturacion consumo'].notna()) & (Tarifas['Suma de facturacion consumo'] != 0),
            'FACTURACION CONSUMO'
        ] = Tarifas['Suma de facturacion consumo']
        Tarifas.loc[Tarifas['tipo de tarifa'].notna(), 'TIPO TARIFA'] = Tarifas['tipo de tarifa']

        Tarifas = Tarifas.drop(columns=['producto', 'Suma de consumo', 'Suma de facturacion consumo', 'tipo de tarifa'])

        problemas = Tarifas[
            Tarifas[['CONSUMO', 'FACTURACION CONSUMO']].isna().any(axis=1) |
            Tarifas[['CONSUMO', 'FACTURACION CONSUMO']].isin([np.inf, -np.inf]).any(axis=1)
        ]
        if not problemas.empty:
            st.error("⚠️ Atención: Se encontraron valores no válidos en las siguientes NIU:")
            st.dataframe(problemas[['NIU', 'CONSUMO', 'FACTURACION CONSUMO']])
            st.stop()
        else:
            Tarifas['CONSUMO'] = np.floor(Tarifas['CONSUMO'] + 0.5).astype(int)
            Tarifas['FACTURACION CONSUMO'] = np.floor(Tarifas['FACTURACION CONSUMO'] + 0.5).astype(int)

        if Tarifas['NIU'].eq('').any():
            st.error("Error: La columna NIU tiene valores vacíos. Revisa TC1 y TC2.")
            st.stop()
        else:
            st.success("✅ Validación exitosa: La columna NIU no tiene valores vacíos.")

        if (Tarifas['CONSUMO'] < 0).any():
            st.error("Error: La columna CONSUMO tiene valores negativos. Verifica los datos.")
            st.stop()
        else:
            st.success("✅ Validación exitosa: La columna CONSUMO no tiene valores negativos.")

        if (Tarifas['FACTURACION CONSUMO'] < 0).any():
            st.error("Error: La columna FACTURACION CONSUMO tiene valores negativos. Verifica los datos.")
            st.stop()
        else:
            st.success("✅ Validación exitosa: La columna FACTURACION CONSUMO no tiene valores negativos.")

        if ((Tarifas['CONSUMO'] == 0) & (Tarifas['FACTURACION CONSUMO'] != 0)).any():
            st.error("Error: Si CONSUMO es 0, FACTURACION CONSUMO debe ser 0.")
            st.stop()
        else:
            st.success("✅ Validación exitosa: No hay inconsistencias entre CONSUMO y FACTURACION CONSUMO.")

        if Tarifas.isnull().any().any():
            st.error("Error: El DataFrame contiene valores nulos. Corrige los datos.")
            st.stop()
        else:
            st.success("✅ Validación exitosa: El DataFrame no tiene valores nulos.")

        # Actualización de la hoja "Consolidado" en el archivo Informe_Tarifas_def
        try:
            file_dict["INFORME_TARIFAS_DEF"].seek(0)
            contenido = file_dict["INFORME_TARIFAS_DEF"].read()
            # Si el archivo es macrohabilitado (.xlsm) se usa keep_vba; de lo contrario no.
            if file_dict["INFORME_TARIFAS_DEF"].name.endswith('.xlsm'):
                tarifas_def = load_workbook(BytesIO(contenido), keep_vba=True)
            else:
                tarifas_def = load_workbook(BytesIO(contenido))
            hoja_consolidado = tarifas_def["Consolidado"]
            if hoja_consolidado.max_row > 1:
                hoja_consolidado.delete_rows(2, hoja_consolidado.max_row - 1)
            data_df = Tarifas.values.tolist()
            for row in data_df:
                hoja_consolidado.append(row)
        except Exception as e:
            st.error(f"Error al actualizar la hoja 'Consolidado': {e}")
            st.stop()

        # Validación Bitácora
        bitacora['Producto'] = bitacora['Producto'].astype(str)
        bitacora = bitacora[bitacora['Tipo Frontera'] == 'Tipo No Regulado']
        ultima_columna_bitacora = bitacora.columns[-1]
        resultado = pd.merge(
            bitacora[['Producto', ultima_columna_bitacora]],
            Tarifas[['NIU', 'CONSUMO']],
            left_on='Producto', right_on='NIU',
            how='left'
        )
        resultado['Diferencia'] = abs(resultado[ultima_columna_bitacora] - resultado['CONSUMO'])
        resultado['Es Diferente'] = resultado['Diferencia'] > 1
        diferencias = resultado[resultado['Es Diferente']][['NIU', 'CONSUMO', ultima_columna_bitacora]]
        st.write("### Tabla de diferencias tarifas con bitacora:")
        st.dataframe(diferencias)
        st.write("### Tabla de Tarifas Generada:")
        st.dataframe(Tarifas)

        # Creación de informe DANE
        informeDane = Tarifas[(Tarifas['UBICACION'] == 'U') & (Tarifas['MUNICIPIO'] == 'POPAYAN')]
        pivot_table = pd.pivot_table(
            informeDane,
            index='ESTRATO',
            values=['NIU', 'CONSUMO', 'FACTURACION CONSUMO'],
            aggfunc={'NIU': 'count', 'CONSUMO': 'sum', 'FACTURACION CONSUMO': 'sum'}
        )
        pivot_table.rename(
            columns={'NIU': 'CONTEO_NIU', 'CONSUMO': 'SUMA_CONSUMO', 'FACTURACION CONSUMO': 'SUMA_FACTURACION'},
            inplace=True
        )
        informeDaneVf = pivot_table.reset_index()
        st.write("### Tabla de informe DANE:")
        st.dataframe(informeDaneVf)

        # Función para descargar informes en un ZIP
        st.write("Descargar los informes")
        def create_zip():
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                excel_buffer = io.BytesIO()
                tarifas_def.save(excel_buffer)
                excel_buffer.seek(0)
                zip_file.writestr("TarifasDefinitivo.xlsx", excel_buffer.getvalue())
                dane_buffer = io.StringIO()
                informeDaneVf.to_csv(dane_buffer, index=False, encoding='utf-8-sig')
                zip_file.writestr("Informe_DANE.csv", dane_buffer.getvalue())
                diferencias_buffer = io.StringIO()
                diferencias.to_csv(diferencias_buffer, index=False, encoding='utf-8-sig')
                zip_file.writestr("Diferencias_Tarifas_Bitacora.csv", diferencias_buffer.getvalue())
            zip_buffer.seek(0)
            return zip_buffer

        st.download_button(
            label="📥 Descargar Tarifas, Informe DANE y Diferencias",
            data=create_zip(),
            file_name="Reportes_Tarifas.zip",
            mime="application/zip"
        )

    else:
        st.error("❌ No se encontraron todas las columnas necesarias en TC1. Verifica el archivo.")

# Botón para limpiar la app
if st.button("Limpiar"):
    st.session_state.clear()
    st.experimental_rerun()
